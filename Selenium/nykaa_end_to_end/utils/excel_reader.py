import os
from openpyxl import load_workbook

class ExcelReader:

    def __init__(self, file_path):
        project_root = os.getcwd()
        self.file_path = os.path.join(project_root, file_path)

    def get_phone(self):
        print("Reading Excel from:", self.file_path)

        wb = load_workbook(self.file_path)
        sheet = wb.active
        return str(sheet["A2"].value)

    def get_shipping_details(self):
        wb = load_workbook(self.file_path)
        sheet = wb.active

        return {
            "phone": str(sheet["A2"].value),
            "pincode": str(sheet["B2"].value),
            "city": sheet["C2"].value,
            "state": sheet["D2"].value,
            "name": sheet["E2"].value,
            "email": sheet["F2"].value,
            "flatno": str(sheet["G2"].value),
            "area": sheet["H2"].value
        }

    def get_card_details(self):
        wb = load_workbook(self.file_path)
        sheet = wb.active

        return {
            "cardno": sheet["I2"].value,
            "date": str(sheet["J2"].value),
            "cvv": str(sheet["K2"].value)
        }
